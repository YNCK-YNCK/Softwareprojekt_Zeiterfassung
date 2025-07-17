import os
import pandas as pd
import openpyxl
from datetime import datetime, date
from models.employee import Employee
from models.employer import Employer
from deutschland.feiertage.api import default_api
from deutschland import feiertage

# Setzen der Standardkonfiguration für die API
configuration = feiertage.Configuration(host="https://feiertage-api.de/api")

def get_german_holidays(year):
    with feiertage.ApiClient(configuration) as api_client:
        api_instance = default_api.DefaultApi(api_client)
        try:
            jahr = str(year)
            nur_land = "NATIONAL"
            nur_daten = 1
            api_response = api_instance.get_feiertage(jahr=jahr, nur_land=nur_land, nur_daten=nur_daten)
            return api_response
        except feiertage.ApiException as e:
            print(f"Exception when calling DefaultApi->get_feiertage: {e}\n")
            return {}

def is_weekend_or_holiday(date):
    if date.weekday() >= 5:  # 5 and 6 correspond to Saturday and Sunday
        return True
    holidays = get_german_holidays(date.year)
    return any(holiday_date == date for holiday_date in holidays.values())

class TimeTrackingController:
    def __init__(self):
        self.employer = Employer(employer_id="001", name="ACME CORPORATION")
        self.tracking = {}  # Dictionary to keep track of who is tracking time
        try:
            self.employees_df = pd.read_excel('employees.xlsx')
            self.worked_hours_df = pd.read_excel('worked_hours.xlsx')
            # Konvertiere die 'date'-Spalte in ein Datum
            self.worked_hours_df['date'] = pd.to_datetime(self.worked_hours_df['date']).dt.date
        except FileNotFoundError:
            self.employees_df = pd.DataFrame(columns=['employee_id', 'name', 'weekly_hours'])
            self.worked_hours_df = pd.DataFrame(columns=['employee_id', 'date', 'hours'])

        for _, row in self.employees_df.iterrows():
            employee = Employee(row['employee_id'], row['name'], row['weekly_hours'])
            self.employer.add_employee(employee)

    def create_employee(self, employee_id, name, weekly_hours):
        employee = Employee(employee_id, name, weekly_hours)
        self.employer.add_employee(employee)
        new_employee = pd.DataFrame([[employee_id, name, weekly_hours]],
                                    columns=['employee_id', 'name', 'weekly_hours'])
        self.employees_df = pd.concat([self.employees_df, new_employee], ignore_index=True)
        self.employees_df.to_excel('employees.xlsx', index=False)
        return employee

    @staticmethod
    def calculate_break_time(total_hours, has_break_logged):
        if total_hours > 9:
            return 15 if has_break_logged else 45
        elif total_hours > 6:
            return 0 if has_break_logged else 30
        return 0

    def start_tracking(self, employee_id):
        if employee_id not in self.tracking:
            self.tracking[employee_id] = datetime.now()
            return True
        return False

    def stop_tracking(self, employee_id):
        if employee_id in self.tracking:
            start_time = self.tracking.pop(employee_id)
            elapsed = datetime.now() - start_time
            self.log_hours_for_employee(employee_id, datetime.now().date(), elapsed.total_seconds() / 3600)
            return True
        return False

    def log_hours_for_employee(self, employee_id, date, hours):
        employee = self.employer.get_employee(employee_id)
        if not employee:
            return False

        date = pd.to_datetime(date)
        existing_hours = self.worked_hours_df[
            (self.worked_hours_df['employee_id'] == employee_id) &
            (self.worked_hours_df['date'] == date)
        ]['hours'].sum()

        total_hours = existing_hours + hours
        event_logger = EventLogger()
        has_break_logged = event_logger.has_break_event(employee_id, date)
        break_time = self.calculate_break_time(total_hours, has_break_logged)
        effective_hours = total_hours - (break_time / 60)

        self.worked_hours_df = self.worked_hours_df[
            ~((self.worked_hours_df['employee_id'] == employee_id) &
              (self.worked_hours_df['date'] == date))
        ]

        new_hours = pd.DataFrame([[employee_id, date, effective_hours]],
                                 columns=['employee_id', 'date', 'hours'])
        self.worked_hours_df = pd.concat([self.worked_hours_df, new_hours], ignore_index=True)
        self.log_break_event(employee_id, date, break_time)
        self.save_to_excel_with_highlights()
        return True

    def save_to_excel_with_highlights(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        headers = ['employee_id', 'date', 'hours']
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
        date_format = 'yyyy-mm-dd'
        for row_num, row in enumerate(self.worked_hours_df.itertuples(index=False), 2):
            for col_num, value in enumerate(row, 1):
                cell = sheet.cell(row=row_num, column=col_num, value=value)
                if col_num == 2:
                    cell.number_format = date_format
            try:
                # Versuchen Sie, das Datum in ein datetime.date-Objekt umzuwandeln
                if isinstance(row.date, str):
                    date_to_check = datetime.strptime(row.date, '%Y-%m-%d').date()
                elif hasattr(row.date, 'date'):  # Für datetime.datetime-Objekte
                    date_to_check = row.date.date()
                elif isinstance(row.date, date):  # Für datetime.date-Objekte
                    date_to_check = row.date
                else:
                    raise ValueError(f"Unsupported date type: {type(row.date)}")

                if is_weekend_or_holiday(date_to_check):
                    for col_num in range(1, len(row) + 1):
                        cell = sheet.cell(row=row_num, column=col_num)
                        cell.fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            except Exception as e:
                print(f"Error processing date {row.date}: {e}")
                continue
        try:
            workbook.save('worked_hours.xlsx')
        except Exception as e:
            print(f"Error saving Excel file: {e}")

    def log_break_event(self, employee_id, date, break_time):
        event_logger = EventLogger()
        event_logger.log_event(employee_id, 'break', f'Pausenzeit von {break_time} Minuten abgezogen.')

    def get_employee_weekly_hours(self, employee_id, year, week):
        weekly_hours = self.worked_hours_df[
            (self.worked_hours_df['employee_id'] == employee_id) &
            (pd.to_datetime(self.worked_hours_df['date']).dt.year == year) &
            (pd.to_datetime(self.worked_hours_df['date']).dt.isocalendar().week == week)
        ]['hours'].sum()
        return weekly_hours

    def get_employee_remaining_hours(self, employee_id, year, week):
        employee = self.employer.get_employee(employee_id)
        if employee:
            weekly_hours_worked = self.get_employee_weekly_hours(employee_id, year, week)
            reduced_hours = employee.get_reduced_hours_for_week(year, week)
            return (employee.weekly_hours - reduced_hours) - weekly_hours_worked
        return None

    def check_warnings(self, employee_id):
        today = datetime.now().date()
        today_hours_series = self.worked_hours_df[
            (self.worked_hours_df['employee_id'] == employee_id) &
            (self.worked_hours_df['date'] == today)
        ]['hours']

        if today_hours_series.empty:
            today_hours = 0
        else:
            # Konvertiere die Spalte in numerische Werte, falls nötig
            today_hours = pd.to_numeric(today_hours_series, errors='coerce').sum()
            # Falls NaN (durch 'coerce'), setze auf 0
            if pd.isna(today_hours):
                today_hours = 0

        print(f"Type of today_hours: {type(today_hours)}")
        print(f"Value of today_hours: {today_hours}")

        if today_hours > 8:
            event_logger = EventLogger()
            event_logger.log_event(employee_id, 'Warning', 'Exceeded 8 hours of work.')
            if today_hours > 11:
                self.stop_tracking(employee_id)
                event_logger.log_event(employee_id, 'System Action', 'Stopped: Exceeded maximum working hours of 11.')
                return "Stopped: Exceeded maximum working hours of 11."
            return "Warning: Exceeded 8 hours of work."
        return None


    def create_employee_directory(self):
        directory_df = self.employees_df[['employee_id', 'name']].copy()
        directory_df.to_excel('employee_directory.xlsx', index=False)

    def create_yearly_summary(self):
        if not os.path.exists('yearly_summaries'):
            os.makedirs('yearly_summaries')

        self.worked_hours_df['date'] = pd.to_datetime(self.worked_hours_df['date'])
        yearly_summaries = self.worked_hours_df.groupby(['employee_id', self.worked_hours_df['date'].dt.year])['hours'].sum().reset_index()

        for employee_id, year, total_hours in yearly_summaries.itertuples(index=False):
            filename = f"yearly_summaries/{int(year)}_employee-{employee_id}.xlsx"
            summary_df = pd.DataFrame([[int(year), total_hours]], columns=['year', 'total_hours'])
            summary_df.to_excel(filename, index=False)

class Message:
    def __init__(self, employee_id, content):
        self.employee_id = employee_id
        self.content = content

class EventLogger:
    def __init__(self, file_path='events.xlsx'):
        self.file_path = file_path
        try:
            self.events_df = pd.read_excel(self.file_path)
        except FileNotFoundError:
            self.events_df = pd.DataFrame(columns=['employee_id', 'date', 'event_type', 'details'])

    def log_event(self, employee_id, event_type, details):
        new_event = pd.DataFrame([[employee_id, datetime.now().strftime("%Y-%m-%d"), event_type, details]],
                                columns=['employee_id', 'date', 'event_type', 'details'])
        self.events_df = pd.concat([self.events_df, new_event], ignore_index=True)
        self.events_df.to_excel(self.file_path, index=False)

    def has_break_event(self, employee_id, date):
        date_str = date.strftime("%Y-%m-%d")
        break_events = self.events_df[
            (self.events_df['employee_id'] == employee_id) &
            (self.events_df['date'] == date_str) &
            (self.events_df['event_type'] == 'break')
        ]
        return not break_events.empty
